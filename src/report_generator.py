"""
Excel Report Generator (Template-Driven)
=========================================
Builds campaign tracker reports from configurable JSON templates.
Each brand can use its own template or the default 12-sheet HBC layout.
Sheets, metrics, columns, styling — everything is customizable.
"""

import os
import logging
import copy
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from . import config
from .report_templates import get_default_template, resolve_metrics

logger = logging.getLogger(__name__)


# ── Style Builder ────────────────────────────────────────────────────

def _build_styles(styling):
    """Build openpyxl style objects from template styling config."""
    primary = styling.get("primary_color", "4472C4")
    dark = styling.get("primary_dark", "2F5496")
    accent = styling.get("accent_color", "70AD47")
    warning = styling.get("warning_color", "FFC000")
    header_bg = styling.get("header_bg", "D6E4F0")
    font_family = styling.get("font_family", "Arial")

    return {
        "blue_fill": PatternFill("solid", fgColor=primary),
        "light_blue_fill": PatternFill("solid", fgColor=header_bg),
        "dark_fill": PatternFill("solid", fgColor=dark),
        "green_fill": PatternFill("solid", fgColor=accent),
        "yellow_fill": PatternFill("solid", fgColor=warning),
        "title_font": Font(name=font_family, bold=True, size=styling.get("title_size", 12), color="FFFFFF"),
        "header_font": Font(name=font_family, bold=True, size=styling.get("header_size", 10)),
        "header_font_w": Font(name=font_family, bold=True, size=styling.get("header_size", 10), color="FFFFFF"),
        "data_font": Font(name=font_family, size=styling.get("data_size", 9)),
        "section_font": Font(name=font_family, bold=True, size=11),
        "thin_border": Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        ),
    }


# ── Helpers ──────────────────────────────────────────────────────────

def _extract_region(name, regions):
    for region in regions:
        if region.lower() in str(name).lower():
            return region
    return "Other"


def _get_week_groups(start_date, end_date):
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    weeks = []
    current = start
    while current <= end:
        days_to_sat = (5 - current.weekday()) % 7
        if days_to_sat == 0 and current != start:
            days_to_sat = 7
        week_end = min(current + timedelta(days=max(days_to_sat, 1) - 1), end)
        month = current.strftime("%b")
        week_num = (current.day - 1) // 7 + 1
        dates = []
        d = current
        while d <= week_end:
            dates.append(d)
            d += timedelta(days=1)
        weeks.append({"label": f"{month} - Week {week_num}", "dates": dates})
        current = week_end + timedelta(days=1)
    return weeks


def _safe_div(a, b, multiplier=1):
    """Safe division returning 0 on error."""
    try:
        if b and b != 0:
            return (a / b) * multiplier
    except:
        pass
    return 0


def _compute_formula(formula, df, spend=0):
    """Compute a formula string against aggregate data.

    Supported variables: spend, reach, impressions, engagement, thruplay,
    clicks, views, completed_views + any df column aggregates.
    """
    if df is None or (hasattr(df, 'empty') and df.empty):
        return 0

    # Build variable context from dataframe
    ctx = {"spend": spend}

    col_map = {
        "reach": "Reach",
        "impressions": ["Impressions", "Impr."],
        "engagement": "Post engagements",
        "thruplay": "ThruPlays",
        "clicks": ["Clicks (all)", "Clicks"],
        "views": "TrueView views",
        "completed_views": ["Completed Views", "Complete Views"],
    }

    for var_name, cols in col_map.items():
        if isinstance(cols, str):
            cols = [cols]
        val = 0
        for col in cols:
            if col in df.columns:
                val = df[col].sum()
                break
        ctx[var_name] = val

    # If spend wasn't passed, try to compute from data
    if spend == 0:
        for col in ["Amount spent (INR)", "Cost"]:
            if col in df.columns:
                ctx["spend"] = df[col].sum()
                break

    # Parse and evaluate formula safely
    try:
        # Replace * with multiplication
        expr = formula.replace("×", "*")
        result = eval(expr, {"__builtins__": {}}, ctx)
        return result if result is not None and not np.isnan(result) and not np.isinf(result) else 0
    except:
        return 0


# ── Sheet Builders ───────────────────────────────────────────────────

def _build_targeted_vs_achieved(ws, meta_df, yt_df, sheet_config, styles, template):
    """Build the Targeted vs Achieved sheet from template config."""
    regions = config.REGIONS
    meta_cfg = config.meta
    yt_cfg = config.google

    # Add region column
    if len(meta_df):
        meta_df = meta_df.copy()
        meta_df["Region"] = meta_df["Campaign name"].apply(lambda x: _extract_region(x, regions))
    if len(yt_df):
        yt_df = yt_df.copy()
        col_name = "Campaign" if "Campaign" in yt_df.columns else yt_df.columns[1] if len(yt_df.columns) > 1 else yt_df.columns[0]
        yt_df["Region"] = yt_df[col_name].apply(lambda x: _extract_region(x, regions))

    platforms = sheet_config.get("platforms", {})

    for platform_key, plat_cfg in platforms.items():
        start_col = plat_cfg.get("start_col", 2)
        label = plat_cfg.get("label", platform_key.upper())
        is_meta = platform_key == "meta"
        df = meta_df if is_meta else yt_df
        cfg = meta_cfg if is_meta else yt_cfg
        t = cfg.targets

        row = 1

        # Platform title
        ws.cell(row=row, column=start_col, value=label).font = styles["title_font"]
        for c in range(start_col, start_col + 5):
            ws.cell(row=row, column=c).fill = styles["blue_fill"]
        row += 1

        # Date info
        ws.cell(row=row, column=start_col, value="Start").font = styles["data_font"]
        ws.cell(row=row, column=start_col + 2, value=cfg.start_date).font = styles["data_font"]
        row += 1
        ws.cell(row=row, column=start_col, value="End").font = styles["data_font"]
        ws.cell(row=row, column=start_col + 2, value=cfg.end_date).font = styles["data_font"]
        row += 1

        try:
            start_dt = datetime.strptime(cfg.start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(cfg.end_date, "%Y-%m-%d")
            ws.cell(row=row, column=start_col, value="Duration").font = styles["data_font"]
            ws.cell(row=row, column=start_col + 2, value=f"{(end_dt - start_dt).days} days").font = styles["data_font"]
        except:
            pass
        row += 2

        # Headers
        headers = ["Metrics", "Target", "Achieved", "%", "Budget Remaining"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=start_col + i, value=h)
            cell.font = styles["header_font"]
            cell.fill = styles["light_blue_fill"]
            cell.border = styles["thin_border"]
        row += 1

        # Volume metrics from template
        spend = 0
        for m in plat_cfg.get("volume_metrics", []):
            m_label = m["label"]
            target_key = m.get("target_key")
            data_col = m.get("data_col")
            default = m.get("default")

            if target_key:
                target = t.get(target_key, cfg.budget if target_key == "amount_spent" else 0)
            else:
                target = default or "--"

            if data_col and data_col in df.columns and len(df):
                agg = m.get("agg", "sum")
                achieved = df[data_col].sum() if agg == "sum" else df[data_col].mean()
            elif "formula" in m:
                achieved = _compute_formula(m["formula"], df, spend)
                target = t.get(target_key, 0) if target_key else "--"
            elif default:
                achieved = default
            else:
                achieved = 0

            if m_label == "Amount Spent":
                spend = achieved

            ws.cell(row=row, column=start_col, value=m_label).font = styles["data_font"]
            ws.cell(row=row, column=start_col + 1, value=target).font = styles["data_font"]
            ws.cell(row=row, column=start_col + 2, value=achieved).font = styles["data_font"]

            if isinstance(target, (int, float)) and target > 0 and isinstance(achieved, (int, float)):
                ws.cell(row=row, column=start_col + 3, value=achieved / target).font = styles["data_font"]
                ws.cell(row=row, column=start_col + 3).number_format = "0.00%"
            else:
                ws.cell(row=row, column=start_col + 3, value="--").font = styles["data_font"]

            if m_label == "Amount Spent" and isinstance(target, (int, float)):
                ws.cell(row=row, column=start_col + 4, value=target - (achieved if isinstance(achieved, (int, float)) else 0)).font = styles["data_font"]
            row += 1

        # Rate metrics
        for m in plat_cfg.get("rate_metrics", []):
            target_key = m.get("target_key")
            target_val = t.get(target_key, 0) if target_key else "--"
            prefix = m.get("prefix", "")
            if prefix and isinstance(target_val, (int, float)):
                display_target = f"{prefix}{target_val}"
            else:
                display_target = target_val

            achieved = _compute_formula(m["formula"], df, spend) if "formula" in m else 0

            ws.cell(row=row, column=start_col, value=m["label"]).font = styles["data_font"]
            ws.cell(row=row, column=start_col + 1, value=display_target).font = styles["data_font"]
            ws.cell(row=row, column=start_col + 2, value=achieved).font = styles["data_font"]
            ws.cell(row=row, column=start_col + 3, value="--").font = styles["data_font"]
            row += 1

        # Extra metrics
        for m in plat_cfg.get("extra_metrics", []):
            if "formula" in m:
                val = _compute_formula(m["formula"], df, spend)
            elif m.get("data_col") and m["data_col"] in df.columns and len(df):
                agg = m.get("agg", "sum")
                val = df[m["data_col"]].sum() if agg == "sum" else df[m["data_col"]].mean()
            else:
                val = m.get("default", 0)

            ws.cell(row=row, column=start_col, value=m["label"]).font = styles["data_font"]
            ws.cell(row=row, column=start_col + 1, value="--").font = styles["data_font"]
            ws.cell(row=row, column=start_col + 2, value=val).font = styles["data_font"]
            row += 1

        # Regional breakdowns
        row += 2
        for region in regions:
            region_df = df[df["Region"] == region] if len(df) and "Region" in df.columns else pd.DataFrame()
            r_budget_meta = meta_cfg.regional_budgets.get(region, 0)
            r_budget_yt = yt_cfg.regional_budgets.get(region, 0)
            r_budget = r_budget_meta if is_meta else r_budget_yt

            ws.cell(row=row, column=start_col, value=f"{label} - {region}").font = styles["section_font"]
            row += 1

            r_headers = ["", "", "Achieved", "%", "Budget Remaining"]
            for i, h in enumerate(r_headers):
                if h:
                    ws.cell(row=row, column=start_col + i, value=h).font = styles["header_font"]
            row += 1

            r_spend = 0
            for m in plat_cfg.get("regional_metrics", []):
                m_label = m["label"]
                data_col = m.get("data_col")
                default = m.get("default")

                if "formula" in m:
                    val = _compute_formula(m["formula"], region_df, r_spend)
                elif data_col and data_col in region_df.columns and len(region_df):
                    agg = m.get("agg", "sum")
                    val = region_df[data_col].sum() if agg == "sum" else region_df[data_col].mean()
                elif default:
                    val = default
                else:
                    val = 0

                if m_label == "Amount Spent":
                    r_spend = val if isinstance(val, (int, float)) else 0

                ws.cell(row=row, column=start_col, value=m_label).font = styles["data_font"]

                if m.get("show_budget"):
                    ws.cell(row=row, column=start_col + 1, value=r_budget).font = styles["data_font"]
                    ws.cell(row=row, column=start_col + 2, value=val).font = styles["data_font"]
                    if r_budget > 0 and isinstance(val, (int, float)):
                        ws.cell(row=row, column=start_col + 3, value=val / r_budget).font = styles["data_font"]
                        ws.cell(row=row, column=start_col + 3).number_format = "0.00%"
                    ws.cell(row=row, column=start_col + 4, value=r_budget - (val if isinstance(val, (int, float)) else 0)).font = styles["data_font"]
                else:
                    ws.cell(row=row, column=start_col + 2, value=val).font = styles["data_font"]

                row += 1
            row += 1

    # Column widths
    for col in range(1, 20):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _build_overview_sheet(ws, df, sheet_config, styles, weeks):
    """Build an overview sheet (campaign/adset/ad) from template."""
    group_col = sheet_config.get("group_col", "Campaign name")
    metrics = sheet_config.get("metrics", [])

    if isinstance(metrics, str):
        # Should have been resolved by resolve_metrics, fall back
        metrics = []

    if df.empty:
        ws.cell(row=1, column=1, value="No data available").font = styles["data_font"]
        return

    # Build date columns
    col = 6
    date_cols = {}
    week_cols = {}

    # Headers
    for hdr, ci in [("Metrics", 2), ("Industry Benchmark", 3), ("Estimated", 4), ("Achieved", 5)]:
        ws.cell(row=1, column=ci, value=hdr).font = styles["header_font_w"]
        ws.cell(row=1, column=ci).fill = styles["dark_fill"]

    for week in weeks:
        week_cols[week["label"]] = col
        cell = ws.cell(row=1, column=col, value=week["label"])
        cell.font = styles["header_font_w"]
        cell.fill = styles["green_fill"]
        col += 1
        for d in week["dates"]:
            date_cols[d.strftime("%Y-%m-%d")] = col
            cell = ws.cell(row=1, column=col, value=d)
            cell.font = styles["data_font"]
            cell.number_format = "YYYY-MM-DD"
            col += 1

    groups = df[group_col].unique() if group_col in df.columns else ["Overall"]
    current_row = 2

    for group in groups:
        gdf = df[df[group_col] == group] if group_col in df.columns else df

        ws.cell(row=current_row, column=1, value=group).font = styles["section_font"]
        ws.cell(row=current_row, column=1).fill = styles["light_blue_fill"]

        for m in metrics:
            label = m.get("label", "")
            data_col = m.get("data_col", "")
            agg_type = m.get("agg", "sum")

            if data_col not in gdf.columns:
                current_row += 1
                continue

            ws.cell(row=current_row, column=2, value=label).font = styles["data_font"]

            # Achieved total
            achieved = gdf[data_col].sum() if agg_type == "sum" else gdf[data_col].mean()
            ws.cell(row=current_row, column=5, value=achieved).font = styles["data_font"]

            # Daily values
            daily = gdf.groupby("Day")[data_col].sum() if agg_type == "sum" else gdf.groupby("Day")[data_col].mean()

            for date_str, c in date_cols.items():
                if date_str in daily.index:
                    ws.cell(row=current_row, column=c, value=daily[date_str]).font = styles["data_font"]

            # Weekly summaries
            for week in weeks:
                wdates = [d.strftime("%Y-%m-%d") for d in week["dates"]]
                wdata = daily[daily.index.isin(wdates)]
                if not wdata.empty:
                    val = wdata.sum() if agg_type == "sum" else wdata.mean()
                    ws.cell(row=current_row, column=week_cols[week["label"]], value=val).font = styles["data_font"]

            current_row += 1
        current_row += 1

    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = 22


def _build_raw_sheet(ws, df, sheet_config, styles):
    """Build a raw data dump sheet, optionally filtering columns."""
    columns = sheet_config.get("columns")

    if df.empty:
        ws.cell(row=1, column=1, value="No data available").font = styles["data_font"]
        return

    # Filter columns if specified in template
    if columns:
        available = [c for c in columns if c in df.columns]
        df = df[available] if available else df

    for ci, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = styles["header_font_w"]
        cell.fill = styles["dark_fill"]
        cell.alignment = Alignment(horizontal="center")

    for ri, row in df.iterrows():
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri + 2, column=ci, value=val).font = styles["data_font"]

    for ci, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), df[col_name].astype(str).map(len).max() if len(df) else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 30)


def _build_creative_perf(ws, df, sheet_config, styles):
    """Build creative performance sheet from template."""
    ws.cell(row=1, column=1, value="Creative Performance Summary").font = styles["section_font"]

    if df.empty:
        ws.cell(row=2, column=1, value="No data available").font = styles["data_font"]
        return

    group_col = sheet_config.get("group_col", "Ad name")
    agg_columns = sheet_config.get("agg_columns", {})
    calc_columns = sheet_config.get("calculated_columns", [])

    if group_col not in df.columns:
        ws.cell(row=2, column=1, value=f"Column '{group_col}' not found in data").font = styles["data_font"]
        return

    # Aggregate
    available_agg = {k: v for k, v in agg_columns.items() if k in df.columns}
    if not available_agg:
        ws.cell(row=2, column=1, value="No matching columns in data").font = styles["data_font"]
        return

    perf = df.groupby(group_col).agg(available_agg).reset_index()

    # Calculated columns
    for calc in calc_columns:
        name = calc["name"]
        formula = calc["formula"]
        try:
            parts = formula.split("/")
            if len(parts) == 2:
                num_col = parts[0].strip()
                den_col = parts[1].strip()
                if num_col in perf.columns and den_col in perf.columns:
                    perf[name] = perf[num_col] / perf[den_col].replace(0, np.nan)
                    perf[name] = perf[name].fillna(0)
        except:
            perf[name] = 0

    # Write headers
    for ci, col_name in enumerate(perf.columns, 1):
        cell = ws.cell(row=2, column=ci, value=col_name)
        cell.font = styles["header_font_w"]
        cell.fill = styles["dark_fill"]

    # Write data
    for ri, row in perf.iterrows():
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri + 3, column=ci, value=val).font = styles["data_font"]


def _build_placeholder_sheet(ws, sheet_config, styles):
    """Build a placeholder/empty sheet."""
    title = sheet_config.get("title", sheet_config.get("name", "Sheet"))
    ws.cell(row=1, column=1, value=title).font = styles["section_font"]
    desc = sheet_config.get("description")
    if desc:
        ws.cell(row=2, column=1, value=desc).font = styles["data_font"]


def _build_custom_sheet(ws, df, sheet_config, styles):
    """Build a fully custom sheet with user-defined columns and formulas."""
    title = sheet_config.get("title", sheet_config.get("name", "Custom"))
    ws.cell(row=1, column=1, value=title).font = styles["section_font"]

    if df is None or df.empty:
        ws.cell(row=2, column=1, value="No data available").font = styles["data_font"]
        return

    columns = sheet_config.get("columns", list(df.columns))
    available = [c for c in columns if c in df.columns]

    if not available:
        ws.cell(row=2, column=1, value="No matching columns").font = styles["data_font"]
        return

    sub_df = df[available]
    _build_raw_sheet(ws, sub_df, {"columns": None}, styles)


# ── Main Generator (Template-Driven) ────────────────────────────────

def generate(meta_data, google_data, output_path=None, template=None):
    """Generate an Excel report using the given template.

    Args:
        meta_data: dict with keys raw_data, campaign_data, adset_data, ad_data
        google_data: dict with keys raw_data, campaign_data, ad_group_data, ad_data
        output_path: where to save the .xlsx file
        template: a resolved template dict (from report_templates.py).
                  If None, uses the default template.
    """
    if template is None:
        template = resolve_metrics(get_default_template())

    if not output_path:
        os.makedirs(config.report.output_dir, exist_ok=True)
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = config.report.filename_pattern.format(date=date_str)
        output_path = os.path.join(config.report.output_dir, filename)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    styles = _build_styles(template.get("styling", {}))
    wb = Workbook()
    first_sheet = True

    # Data lookups by platform
    platform_data = {
        "meta": {
            "raw_data": meta_data.get("raw_data", pd.DataFrame()),
            "campaign_data": meta_data.get("campaign_data", pd.DataFrame()),
            "adset_data": meta_data.get("adset_data", pd.DataFrame()),
            "ad_data": meta_data.get("ad_data", pd.DataFrame()),
        },
        "youtube": {
            "raw_data": google_data.get("raw_data", pd.DataFrame()),
            "campaign_data": google_data.get("campaign_data", pd.DataFrame()),
            "ad_group_data": google_data.get("ad_group_data", pd.DataFrame()),
            "ad_data": google_data.get("ad_data", pd.DataFrame()),
        }
    }

    # Data key to platform data mapping for overview sheets
    data_key_map = {
        "meta": {
            "Campaign name": "campaign_data",
            "Ad set name": "adset_data",
            "Ad name": "ad_data",
        },
        "youtube": {
            "Campaign": "campaign_data",
            "Ad group": "ad_group_data",
            "Ad name": "ad_data",
        }
    }

    sheets = template.get("sheets", [])

    for sheet_cfg in sheets:
        if not sheet_cfg.get("enabled", True):
            continue

        sheet_name = sheet_cfg.get("name", "Sheet")
        sheet_type = sheet_cfg.get("type", "raw")

        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(sheet_name)

        try:
            if sheet_type == "summary":
                _build_targeted_vs_achieved(
                    ws, meta_data["raw_data"], google_data["raw_data"],
                    sheet_cfg, styles, template
                )

            elif sheet_type == "overview":
                platform = sheet_cfg.get("platform", "meta")
                group_col = sheet_cfg.get("group_col", "Campaign name")

                # Determine which data to use
                data_key = data_key_map.get(platform, {}).get(group_col, "campaign_data")
                # Also allow explicit data_key override
                data_key = sheet_cfg.get("data_key", data_key)
                df = platform_data.get(platform, {}).get(data_key, pd.DataFrame())

                # Get weeks for this platform
                cfg = config.meta if platform == "meta" else config.google
                weeks = _get_week_groups(cfg.start_date, cfg.end_date)

                _build_overview_sheet(ws, df, sheet_cfg, styles, weeks)

            elif sheet_type == "raw":
                platform = sheet_cfg.get("platform", "meta")
                data_key = sheet_cfg.get("data_key", "raw_data")
                df = platform_data.get(platform, {}).get(data_key, pd.DataFrame())
                _build_raw_sheet(ws, df, sheet_cfg, styles)

            elif sheet_type == "creative":
                platform = sheet_cfg.get("platform", "meta")
                df = platform_data.get(platform, {}).get("ad_data", pd.DataFrame())
                _build_creative_perf(ws, df, sheet_cfg, styles)

            elif sheet_type == "placeholder":
                _build_placeholder_sheet(ws, sheet_cfg, styles)

            elif sheet_type == "custom":
                platform = sheet_cfg.get("platform", "meta")
                data_key = sheet_cfg.get("data_key", "raw_data")
                df = platform_data.get(platform, {}).get(data_key, pd.DataFrame())
                _build_custom_sheet(ws, df, sheet_cfg, styles)

            else:
                _build_placeholder_sheet(ws, sheet_cfg, styles)

        except Exception as e:
            logger.error(f"Error building sheet '{sheet_name}': {e}")
            ws.cell(row=1, column=1, value=f"Error: {str(e)}").font = styles["data_font"]

    # Build any custom_sheets defined in template
    for custom in template.get("custom_sheets", []):
        if not custom.get("enabled", True):
            continue
        ws = wb.create_sheet(custom.get("name", "Custom"))
        platform = custom.get("platform", "meta")
        data_key = custom.get("data_key", "raw_data")
        df = platform_data.get(platform, {}).get(data_key, pd.DataFrame())
        _build_custom_sheet(ws, df, custom, styles)

    wb.save(output_path)
    logger.info(f"Report saved: {output_path}")
    return output_path
